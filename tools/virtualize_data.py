"""
礼簿数据虚拟化脚本
用于生成示例数据，替换真实姓名和金额
"""
import random
from openpyxl import load_workbook, Workbook
import json
import os

# 常见姓氏
SURNAMES = ['张', '王', '李', '刘', '陈', '杨', '黄', '赵', '周', '吴',
            '徐', '孙', '马', '朱', '胡', '郭', '何', '林', '罗', '梁',
            '宋', '郑', '谢', '韩', '唐', '冯', '于', '董', '萧', '程']

# 常见单字名（用于2个字的名字，占20%）
# 男性单字名
MALE_SINGLE_NAMES = ['伟', '强', '磊', '军', '洋', '勇', '杰', '涛', '明', '超',
                     '刚', '浩', '鹏', '辉', '帅', '建', '峰', '宇', '文', '斌', '鑫', '波', '龙', '飞']

# 女性单字名
FEMALE_SINGLE_NAMES = ['芳', '娜', '敏', '静', '丽', '艳', '霞', '婷', '玲', '红',
                       '慧', '雪', '梅', '兰', '英', '琳', '莉', '萍', '颖', '倩']

# 常见双字名（用于3个字的名字，占80%）
# 男性双字名
MALE_DOUBLE_NAMES = ['建国', '国强', '志强', '思远', '国栋', '建华', '军伟', '志勇',
                     '文博', '子轩', '浩然', '天宇', '俊杰', '宇航', '博文', '明轩',
                     '志明', '海涛', '晓东', '伟强', '国庆', '建军', '树林', '金龙']

# 女性双字名
FEMALE_DOUBLE_NAMES = ['秀英', '秀兰', '桂英', '小红', '小芳', '文静', '雨涵', '梦琪',
                       '嘉怡', '雨婷', '诗涵', '美玲', '晓燕', '丽华', '淑芬', '春梅',
                       '海燕', '玉兰', '桂兰', '秀珍', '欣怡', '婉婷', '雅琪', '诗韵']

# 备注（括号内容）- 按性别和唯一性分类

# 男性唯一性亲戚（只能有一个）
MALE_UNIQUE_NOTES = ['爸爸', '大伯', '二伯', '三伯', '姑父', '岳父', '公公']

# 男性可重复亲戚（可以有多个）
MALE_REPEATABLE_NOTES = ['叔叔', '表哥', '堂弟', '哥哥', '弟弟', '舅舅',
                         '姐夫', '妹夫', '姨父', '伯伯', '叔父', '堂哥', '堂弟']

# 女性唯一性亲戚（只能有一个）
FEMALE_UNIQUE_NOTES = ['妈妈', '大姨', '二姨', '三姨', '姑姑', '岳母', '婆婆']

# 女性可重复亲戚（可以有多个）
FEMALE_REPEATABLE_NOTES = ['阿姨', '表妹', '堂姐', '姐姐', '妹妹', '姨妈',
                           '婶婶', '舅妈', '伯母', '姑母', '表姐', '堂妹', '小姨']

# 中性关系（男女通用，可以有多个）
NEUTRAL_NOTES = ['同学', '同事', '朋友', '邻居', '亲戚', '老师', '战友', '发小']

# 支付方式权重
PAYMENT_METHODS = {
    '微信': 0.5,      # 50%
    '支付宝': 0.25,   # 25%
    '现金': 0.2,      # 20%
    '未记录': 0.05    # 5%
}

# 金额范围（根据常见礼金金额）
AMOUNT_RANGES = [
    (100, 200, 0.15),   # 15% 概率 100-200元
    (200, 500, 0.30),   # 30% 概率 200-500元
    (500, 1000, 0.35),  # 35% 概率 500-1000元
    (1000, 2000, 0.15), # 15% 概率 1000-2000元
    (2000, 5000, 0.05), # 5% 概率 2000-5000元
]

# 常见整数金额（吉利数字）
COMMON_AMOUNTS = [
    100, 200, 300, 500, 600, 800, 1000, 1200, 1500, 1800, 2000, 2800, 3000, 5000
]

def generate_chinese_name(add_note=False, used_unique_notes=None):
    """生成随机中文姓名
    80%概率生成3个字的名字（姓+双字名）
    20%概率生成2个字的名字（姓+单字名）
    根据性别匹配合适的备注关系，避免唯一性亲戚重复

    Args:
        add_note: 是否添加备注
        used_unique_notes: 已使用的唯一性备注集合（避免重复）
    """
    if used_unique_notes is None:
        used_unique_notes = set()

    surname = random.choice(SURNAMES)

    # 随机选择性别
    is_male = random.random() < 0.5  # 50% 男性，50% 女性

    # 80% 概率使用双字名（3个字），20% 概率使用单字名（2个字）
    if random.random() < 0.8:
        # 双字名
        if is_male:
            name = random.choice(MALE_DOUBLE_NAMES)
        else:
            name = random.choice(FEMALE_DOUBLE_NAMES)
    else:
        # 单字名
        if is_male:
            name = random.choice(MALE_SINGLE_NAMES)
        else:
            name = random.choice(FEMALE_SINGLE_NAMES)

    full_name = surname + name

    # 30% 概率添加备注，根据性别选择合适的备注
    if add_note and random.random() < 0.3:
        note = None

        # 选择备注：70%使用性别匹配的亲戚关系，30%使用中性关系
        if random.random() < 0.7:
            # 使用性别匹配的亲戚关系
            if is_male:
                # 合并男性唯一性和可重复备注
                unique_notes = [n for n in MALE_UNIQUE_NOTES if n not in used_unique_notes]
                repeatable_notes = MALE_REPEATABLE_NOTES

                # 50%概率尝试使用唯一性备注（如果还有可用的）
                if unique_notes and random.random() < 0.5:
                    note = random.choice(unique_notes)
                    used_unique_notes.add(note)
                else:
                    note = random.choice(repeatable_notes)
            else:
                # 女性
                unique_notes = [n for n in FEMALE_UNIQUE_NOTES if n not in used_unique_notes]
                repeatable_notes = FEMALE_REPEATABLE_NOTES

                # 50%概率尝试使用唯一性备注（如果还有可用的）
                if unique_notes and random.random() < 0.5:
                    note = random.choice(unique_notes)
                    used_unique_notes.add(note)
                else:
                    note = random.choice(repeatable_notes)
        else:
            # 使用中性关系（可重复）
            note = random.choice(NEUTRAL_NOTES)

        if note:
            full_name += f'({note})'

    return full_name

def generate_amount():
    """生成随机金额（所有金额都是100元的整数倍）"""
    # 70% 概率使用常见整数金额
    if random.random() < 0.7:
        return float(random.choice(COMMON_AMOUNTS))

    # 30% 概率使用范围随机金额
    ranges_and_probs = [(r, p) for r, p in [(AMOUNT_RANGES[i], AMOUNT_RANGES[i][2])
                                              for i in range(len(AMOUNT_RANGES))]]
    rand = random.random()
    cumulative = 0

    for (min_amt, max_amt, _), prob in ranges_and_probs:
        cumulative += prob
        if rand <= cumulative:
            # 生成100的倍数（百元整）
            base_amount = random.randint(min_amt // 100, max_amt // 100) * 100
            return float(base_amount)

    return 500.0  # 默认值

def generate_payment_method():
    """根据权重生成支付方式"""
    rand = random.random()
    cumulative = 0

    for method, prob in PAYMENT_METHODS.items():
        cumulative += prob
        if rand <= cumulative:
            return method

    return '微信'  # 默认值

def number_to_chinese(num):
    """将数字转换为中文大写（简化版）"""
    if num == 0:
        return '零元整'

    digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    units = ['', '拾', '佰', '仟']
    big_units = ['', '万', '亿']

    parts = f"{num:.2f}".split('.')
    integer_part = int(parts[0])
    decimal_part = parts[1]

    if integer_part == 0:
        jiao = int(decimal_part[0])
        fen = int(decimal_part[1])
        result = ''
        if jiao > 0:
            result += digits[jiao] + '角'
        if fen > 0:
            result += digits[fen] + '分'
        return result or '零元整'

    result = ''
    unit_index = 0
    zero_count = 0

    while integer_part > 0:
        section = integer_part % 10000
        if section != 0:
            section_str = ''
            section_num = section
            pos = 0

            while section_num > 0:
                digit = section_num % 10
                if digit == 0:
                    zero_count += 1
                else:
                    if zero_count > 0:
                        section_str = '零' + section_str
                        zero_count = 0
                    section_str = digits[digit] + units[pos] + section_str
                section_num = section_num // 10
                pos += 1

            result = section_str + big_units[unit_index] + result
        elif result != '':
            zero_count += 1

        integer_part = integer_part // 10000
        unit_index += 1

    if result.endswith('零'):
        result = result[:-1]

    result += '元'

    jiao = int(decimal_part[0])
    fen = int(decimal_part[1])

    if jiao == 0 and fen == 0:
        result += '整'
    else:
        if jiao > 0:
            result += digits[jiao] + '角'
        if fen > 0:
            if jiao == 0:
                result += '零'
            result += digits[fen] + '分'

    return result

def virtualize_excel(input_file, output_file, num_records=50):
    """虚拟化Excel文件"""
    print(f"正在生成 {num_records} 条虚拟数据...")

    wb = Workbook()
    ws = wb.active
    ws.title = '汇总'

    # 表头
    headers = ['姓名', '礼金/元', '礼金/大写', '支付方式']
    ws.append(headers)

    # 生成虚拟数据
    total_amount = 0
    used_unique_notes = set()  # 跟踪已使用的唯一性备注

    for i in range(num_records):
        name = generate_chinese_name(add_note=True, used_unique_notes=used_unique_notes)
        amount = generate_amount()
        amount_chinese = number_to_chinese(amount)
        payment_method = generate_payment_method()

        ws.append([name, amount, amount_chinese, payment_method])
        total_amount += amount

        if (i + 1) % 10 == 0:
            print(f"已生成 {i + 1} 条记录...")

    # 添加总计行
    ws.append([])
    ws.append([
        '总计',
        total_amount,
        number_to_chinese(total_amount),
        ''
    ])

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12

    # 创建附件工作表
    ws2 = wb.create_sheet('附件')
    ws2.append(['支付方式'])
    ws2.append(['微信'])
    ws2.append(['现金'])
    ws2.append(['支付宝'])
    ws2.append(['未记录'])

    # 保存文件
    wb.save(output_file)
    print(f"\n✓ 虚拟数据已保存到: {output_file}")
    print(f"✓ 总记录数: {num_records}")
    print(f"✓ 总金额: {total_amount:.2f} 元")

def virtualize_json(input_file, output_file, num_records=50):
    """虚拟化JSON文件"""
    print(f"正在生成 {num_records} 条虚拟数据...")

    records = []
    used_unique_notes = set()  # 跟踪已使用的唯一性备注

    for i in range(num_records):
        name = generate_chinese_name(add_note=True, used_unique_notes=used_unique_notes)
        amount = generate_amount()
        payment_method = generate_payment_method()

        record = {
            'id': 1000000000000 + i,
            'name': name,
            'amount': amount,
            'amountChinese': number_to_chinese(amount),
            'paymentMethod': payment_method,
            'timestamp': f'2025-12-{random.randint(20, 31):02d}T{random.randint(8, 20):02d}:{random.randint(0, 59):02d}:00.000Z'
        }
        records.append(record)

        if (i + 1) % 10 == 0:
            print(f"已生成 {i + 1} 条记录...")

    # 保存JSON
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    total_amount = sum(r['amount'] for r in records)
    print(f"\n✓ 虚拟数据已保存到: {output_file}")
    print(f"✓ 总记录数: {num_records}")
    print(f"✓ 总金额: {total_amount:.2f} 元")

def main():
    print("=" * 60)
    print("礼簿数据虚拟化工具")
    print("=" * 60)
    print()

    # 获取项目根目录（src的父目录）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    data_dir = os.path.join(project_root, 'data')

    # 创建data文件夹（如果不存在）
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
        print(f"✓ 创建data文件夹: {data_dir}")
        print()

    # 生成虚拟数据
    num_records = int(input("请输入要生成的记录数（默认50）: ") or "50")

    # 生成Excel
    output_excel = os.path.join(data_dir, '礼簿_示例数据.xlsx')
    virtualize_excel(None, output_excel, num_records)

    print()

    # 生成JSON
    output_json = os.path.join(data_dir, 'data_示例数据.json')
    virtualize_json(None, output_json, num_records)

    print()
    print("=" * 60)
    print("虚拟化完成！")
    print("=" * 60)
    print()
    print("生成的文件：")
    print(f"  1. {output_excel}")
    print(f"  2. {output_json}")
    print()
    print("这些文件可以安全地用于：")
    print("  - 项目展示")
    print("  - 功能演示")
    print("  - 截图说明")
    print("  - 公开发布")

if __name__ == '__main__':
    main()
