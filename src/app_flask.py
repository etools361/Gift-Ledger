from flask import Flask, render_template, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import json
from datetime import datetime

# 获取项目根目录（src的父目录）
script_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(script_dir)

app = Flask(__name__,
            template_folder=os.path.join(project_root, 'templates'),
            static_folder=os.path.join(project_root, 'static'))

# 配置
EXCEL_FILE = os.path.join(project_root, 'data', '礼簿.xlsx')
DATA_FILE = os.path.join(project_root, 'data', 'data.json')

def number_to_chinese(num):
    """将数字转换为中文大写"""
    if num == 0:
        return '零元整'

    digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    units = ['', '拾', '佰', '仟']
    big_units = ['', '万', '亿']

    # 分离整数和小数部分
    num_str = f"{num:.2f}"
    parts = num_str.split('.')
    integer_part = int(parts[0])
    decimal_part = parts[1]

    if integer_part == 0:
        # 处理小数部分
        jiao = int(decimal_part[0])
        fen = int(decimal_part[1])
        result = ''
        if jiao > 0:
            result += digits[jiao] + '角'
        if fen > 0:
            result += digits[fen] + '分'
        return result or '零元整'

    # 处理整数部分
    result = ''
    unit_index = 0
    zero_count = 0

    while integer_part > 0:
        section = integer_part % 10000
        if section != 0:
            section_str = ''
            section_num = section
            pos = 0

            while section_num > 0:
                digit = section_num % 10
                if digit == 0:
                    zero_count += 1
                else:
                    if zero_count > 0:
                        section_str = '零' + section_str
                        zero_count = 0
                    section_str = digits[digit] + units[pos] + section_str
                section_num = section_num // 10
                pos += 1

            result = section_str + big_units[unit_index] + result
        elif result != '':
            zero_count += 1

        integer_part = integer_part // 10000
        unit_index += 1

    # 去掉末尾的"零"
    if result.endswith('零'):
        result = result[:-1]

    result += '元'

    # 处理小数部分
    jiao = int(decimal_part[0])
    fen = int(decimal_part[1])

    if jiao == 0 and fen == 0:
        result += '整'
    else:
        if jiao > 0:
            result += digits[jiao] + '角'
        if fen > 0:
            if jiao == 0:
                result += '零'
            result += digits[fen] + '分'

    return result

def load_from_excel():
    """从Excel文件加载数据"""
    if not os.path.exists(EXCEL_FILE):
        return None

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb['汇总']

        records = []
        # 跳过表头，从第2行开始读取
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过空行和总计行
            if not row[0] or row[0] == '总计':
                break

            name = row[0] if row[0] else ''
            amount = float(row[1]) if row[1] else 0.0
            amount_chinese = row[2] if row[2] else number_to_chinese(amount)
            payment_method = row[3] if row[3] else '未记录'

            # 生成唯一ID（基于时间戳）
            record_id = int(datetime.now().timestamp() * 1000) + len(records)

            record = {
                'id': record_id,
                'name': name,
                'amount': amount,
                'amountChinese': amount_chinese,
                'paymentMethod': payment_method,
                'timestamp': datetime.now().isoformat() + 'Z'
            }
            records.append(record)

        wb.close()
        return records
    except Exception as e:
        print(f"从Excel加载数据失败: {e}")
        return None

def load_data():
    """从文件加载数据，优先从Excel加载"""
    # 优先尝试从Excel加载
    excel_data = load_from_excel()
    if excel_data is not None:
        print(f"✓ 从 {EXCEL_FILE} 加载了 {len(excel_data)} 条记录")
        # 同步保存到JSON
        save_data(excel_data)
        return excel_data

    # 如果Excel不存在或加载失败，尝试从JSON加载
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print(f"✓ 从 {DATA_FILE} 加载了 {len(data)} 条记录")
            return data

    print("! 未找到数据文件，返回空数据")
    return []

def save_data(records):
    """保存数据到JSON文件"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def save_to_excel(records):
    """保存数据到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = '汇总'

    # 表头
    headers = ['姓名', '礼金/元', '礼金/大写', '支付方式']
    ws.append(headers)

    # 设置表头样式
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 添加数据
    for record in records:
        ws.append([
            record['name'],
            record['amount'],
            record['amountChinese'],
            record['paymentMethod']
        ])

    # 添加总计行
    total_amount = sum(r['amount'] for r in records)
    ws.append([])
    ws.append([
        '总计',
        total_amount,
        number_to_chinese(total_amount),
        ''
    ])

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12

    # 创建附件工作表
    ws2 = wb.create_sheet('附件')
    ws2.append(['支付方式'])
    ws2.append(['微信'])
    ws2.append(['现金'])
    ws2.append(['支付宝'])
    ws2.append(['未记录'])

    # 保存文件
    wb.save(EXCEL_FILE)

def load_from_excel():
    """从Excel文件加载数据"""
    if not os.path.exists(EXCEL_FILE):
        return []

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[0] == '总计':
            continue

        name = str(row[0]) if row[0] else ''
        amount = float(row[1]) if row[1] else 0
        payment_method = str(row[3]) if row[3] else '未记录'

        if name and amount > 0:
            records.append({
                'id': int(datetime.now().timestamp() * 1000),
                'name': name,
                'amount': amount,
                'amountChinese': number_to_chinese(amount),
                'paymentMethod': payment_method,
                'timestamp': datetime.now().isoformat()
            })

    return records

@app.route('/')
def index():
    """渲染主页"""
    return render_template('index.html')

@app.route('/api/records', methods=['GET'])
def get_records():
    """获取所有记录"""
    records = load_data()
    return jsonify(records)

@app.route('/api/records', methods=['POST'])
def save_records():
    """保存所有记录"""
    records = request.json
    save_data(records)
    save_to_excel(records)  # 自动保存到Excel
    return jsonify({'success': True, 'message': '保存成功'})

@app.route('/api/excel/load', methods=['GET'])
def load_excel():
    """从Excel加载数据"""
    try:
        records = load_from_excel()
        save_data(records)  # 同步到JSON
        return jsonify({'success': True, 'records': records})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/excel/export', methods=['POST'])
def export_excel():
    """导出Excel文件"""
    try:
        records = request.json
        save_to_excel(records)
        return send_file(
            EXCEL_FILE,
            as_attachment=True,
            download_name='礼簿_导出.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/import', methods=['POST'])
def import_data():
    """导入数据"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '没有上传文件'}), 400

        file = request.files['file']
        filename = file.filename.lower()

        if filename.endswith('.json'):
            # 导入JSON
            data = json.load(file)
            records = []
            for record in data:
                records.append({
                    'id': record.get('id', int(datetime.now().timestamp() * 1000)),
                    'name': record['name'],
                    'amount': record['amount'],
                    'amountChinese': number_to_chinese(record['amount']),
                    'paymentMethod': record.get('paymentMethod', '未记录'),
                    'timestamp': record.get('timestamp', datetime.now().isoformat())
                })
        elif filename.endswith(('.xlsx', '.xls')):
            # 导入Excel
            file.save('temp_import.xlsx')
            wb = load_workbook('temp_import.xlsx')
            ws = wb.active

            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or row[0] == '总计':
                    continue

                name = str(row[0]) if row[0] else ''
                amount = float(row[1] or row[2] or 0)
                payment_method = str(row[3] if len(row) > 3 and row[3] else '未记录')

                if name and amount > 0:
                    records.append({
                        'id': int(datetime.now().timestamp() * 1000),
                        'name': name,
                        'amount': amount,
                        'amountChinese': number_to_chinese(amount),
                        'paymentMethod': payment_method,
                        'timestamp': datetime.now().isoformat()
                    })

            os.remove('temp_import.xlsx')
        else:
            return jsonify({'success': False, 'message': '不支持的文件格式'}), 400

        save_data(records)
        save_to_excel(records)
        return jsonify({'success': True, 'records': records})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

if __name__ == '__main__':
    # 创建data文件夹（如果不存在）
    data_dir = os.path.dirname(DATA_FILE)
    if data_dir and not os.path.exists(data_dir):
        os.makedirs(data_dir)
        print(f"✓ 创建data文件夹")

    # 初始化数据
    if not os.path.exists(DATA_FILE):
        if os.path.exists(EXCEL_FILE):
            # 从Excel加载初始数据
            records = load_from_excel()
            save_data(records)
        else:
            # 创建空数据文件
            save_data([])

    app.run(debug=True, host='0.0.0.0', port=5000)
